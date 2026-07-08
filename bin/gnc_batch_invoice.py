#!/usr/bin/env python3
"""
GnuCash Invoice Batch Creator
Copyright (C) 2026 Alan Johnson / contributors
SPDX-License-Identifier: GPL-3.0-or-later

Read a GnuCash SQLite/XML book for customer and invoice IDs, scan uploaded
customer lists, scan selectable customers from the book, and generate GnuCash invoice import CSV rows.
"""
from __future__ import annotations

import argparse
import base64
import csv
import gzip
import hashlib
import json
import os
import re
import shutil
import secrets
import socket
import sqlite3
import subprocess
import struct
import tempfile
import time
import sys
import zipfile
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

# Match the PHP application runtime policy: files created by this helper should
# be group-readable/writable in setgid profile directories. Ownership still
# follows the PHP-FPM user that launched this script.
os.umask(0o007)


def apply_runtime_permissions(path: Path) -> None:
    try:
        if path.is_dir():
            path.chmod(0o2770)
        elif path.exists():
            path.chmod(0o660)
    except OSError:
        pass


def ensure_runtime_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)
    apply_runtime_permissions(path)


FIELDNAMES = [
    "id",
    "date_opened",
    "owner_id",
    "billingid",
    "notes",
    "date",
    "desc",
    "action",
    "account",
    "quantity",
    "price",
    "disc_type",
    "disc_how",
    "discount",
    "taxable",
    "taxincluded",
    "tax_table",
    "date_posted",
    "due_date",
    "account_posted",
    "memo_posted",
    "accu_splits",
]

CUSTOMER_HEADER_NAMES = {
    "customer_id",
    "customer id",
    "customer",
    "gnucash_customer_id",
    "gnucash customer id",
    "owner_id",
    "owner id",
    "id",
}


@dataclass
class Customer:
    id: str
    name: str = ""
    active: str = ""
    guid: str = ""
    billing_name: str = ""


def print_json(data: Any) -> None:
    json.dump(data, sys.stdout, indent=2, ensure_ascii=False)
    sys.stdout.write("\n")


def err(message: str, code: int = 1) -> None:
    print_json({"ok": False, "error": message})
    raise SystemExit(code)


def text_value(value: Any) -> str:
    return "" if value is None else str(value)


def local_name(tag: str) -> str:
    if "}" in tag:
        return tag.rsplit("}", 1)[1]
    if ":" in tag:
        return tag.rsplit(":", 1)[1]
    return tag


def child_text(element: ET.Element, wanted: str) -> str:
    for child in list(element):
        if local_name(child.tag).lower() == wanted.lower():
            return (child.text or "").strip()
    return ""


def try_sqlite_book(book_path: Path) -> tuple[list[Customer], list[str]] | None:
    try:
        con = sqlite3.connect(f"file:{book_path}?mode=ro", uri=True)
    except sqlite3.Error:
        return None
    try:
        cur = con.cursor()
        tables = {row[0] for row in cur.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        if "customers" not in tables and "invoices" not in tables:
            return None

        customers: list[Customer] = []
        if "customers" in tables:
            columns = {row[1] for row in cur.execute("PRAGMA table_info(customers)")}
            id_col = "id" if "id" in columns else None
            if id_col:
                name_col = "name" if "name" in columns else "''"
                active_col = "active" if "active" in columns else "''"
                guid_col = "guid" if "guid" in columns else "''"
                billing_col = "addr_name" if "addr_name" in columns else "''"
                query = f"SELECT {id_col}, {name_col}, {active_col}, {guid_col}, {billing_col} FROM customers ORDER BY {id_col}"
                for row in cur.execute(query):
                    cid = str(row[0] or "").strip()
                    if cid:
                        customers.append(Customer(cid, text_value(row[1]), text_value(row[2]), text_value(row[3]), text_value(row[4])))

        invoice_ids: list[str] = []
        if "invoices" in tables:
            columns = {row[1] for row in cur.execute("PRAGMA table_info(invoices)")}
            if "id" in columns:
                for row in cur.execute("SELECT id FROM invoices WHERE id IS NOT NULL"):
                    iid = str(row[0] or "").strip()
                    if iid:
                        invoice_ids.append(iid)
        return customers, invoice_ids
    except sqlite3.Error as exc:
        raise RuntimeError(f"SQLite book scan failed: {exc}") from exc
    finally:
        con.close()




def try_sqlite_accounts(book_path: Path) -> list[dict[str, Any]] | None:
    """Return GnuCash account metadata from a SQLite book.

    The invoice importer expects account names/full account paths, not GUIDs,
    so the web UI uses these rows to offer valid account selections.
    """
    try:
        con = sqlite3.connect(f"file:{book_path}?mode=ro", uri=True)
        con.row_factory = sqlite3.Row
    except sqlite3.Error:
        return None
    try:
        cur = con.cursor()
        tables = {row[0] for row in cur.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        if "accounts" not in tables:
            return None
        rows = {str(row["guid"]): dict(row) for row in cur.execute("SELECT guid, name, account_type, parent_guid FROM accounts")}

        def full_name(guid: str, seen: set[str] | None = None) -> str:
            seen = seen or set()
            row = rows.get(guid)
            if not row:
                return guid
            name = str(row.get("name") or "")
            parent = str(row.get("parent_guid") or "")
            if not parent or parent in seen or parent not in rows:
                return name
            parent_name = full_name(parent, seen | {guid})
            if parent_name in {"Root Account", "Root", ""}:
                return name
            return parent_name + ":" + name

        accounts = []
        for guid, row in rows.items():
            account_type = str(row.get("account_type") or "")
            name = str(row.get("name") or "")
            if not name:
                continue
            accounts.append({
                "guid": guid,
                "name": name,
                "full_name": full_name(guid),
                "account_type": account_type,
            })
        accounts.sort(key=lambda a: natural_key(str(a.get("full_name") or a.get("name") or "")))
        return accounts
    finally:
        con.close()


def scan_xml_accounts(book_path: Path) -> list[dict[str, Any]]:
    try:
        root = ET.fromstring(open_xml_bytes(book_path))
    except Exception:
        return []
    rows: dict[str, dict[str, str]] = {}
    for elem in root.iter():
        if local_name(elem.tag) == "GncAccount":
            guid = child_text(elem, "id") or child_text(elem, "guid")
            name = child_text(elem, "name")
            account_type = child_text(elem, "type")
            parent = child_text(elem, "parent")
            if guid and name:
                rows[guid] = {"guid": guid, "name": name, "account_type": account_type, "parent_guid": parent}

    def full_name(guid: str, seen: set[str] | None = None) -> str:
        seen = seen or set()
        row = rows.get(guid)
        if not row:
            return guid
        name = row.get("name", "")
        parent = row.get("parent_guid", "")
        if not parent or parent in seen or parent not in rows:
            return name
        parent_name = full_name(parent, seen | {guid})
        if parent_name in {"Root Account", "Root", ""}:
            return name
        return parent_name + ":" + name

    accounts = [{"guid": guid, "name": row["name"], "full_name": full_name(guid), "account_type": row.get("account_type", "")} for guid, row in rows.items()]
    accounts.sort(key=lambda a: natural_key(str(a.get("full_name") or a.get("name") or "")))
    return accounts


def scan_accounts(book_path: str | Path) -> list[dict[str, Any]]:
    path = Path(book_path).expanduser()
    sqlite_accounts = try_sqlite_accounts(path)
    if sqlite_accounts is not None:
        return sqlite_accounts
    return scan_xml_accounts(path)


def is_income_account(account: dict[str, Any]) -> bool:
    return str(account.get("account_type") or "").upper() == "INCOME"


def is_receivable_account(account: dict[str, Any]) -> bool:
    return str(account.get("account_type") or "").upper() in {"RECEIVABLE", "A/RECEIVABLE"}

def open_xml_bytes(book_path: Path) -> bytes:
    raw = book_path.read_bytes()
    if raw[:2] == b"\x1f\x8b":
        return gzip.decompress(raw)
    return raw


def scan_xml_book(book_path: Path) -> tuple[list[Customer], list[str]]:
    try:
        root = ET.fromstring(open_xml_bytes(book_path))
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(f"Unable to parse book as SQLite or XML/gzipped XML: {exc}") from exc

    customers: list[Customer] = []
    invoice_ids: list[str] = []

    for elem in root.iter():
        lname = local_name(elem.tag)
        if lname == "GncCustomer":
            cid = child_text(elem, "id")
            if cid:
                customers.append(
                    Customer(
                        id=cid,
                        name=child_text(elem, "name"),
                        active=child_text(elem, "active"),
                        guid=child_text(elem, "guid"),
                        billing_name=child_text(elem, "addr-name") or child_text(elem, "billing_name"),
                    )
                )
        elif lname == "GncInvoice":
            iid = child_text(elem, "id")
            if iid:
                invoice_ids.append(iid)

    customers.sort(key=lambda c: natural_key(c.id))
    return customers, invoice_ids


def scan_book(book_path: str | Path) -> tuple[list[Customer], list[str]]:
    path = Path(book_path).expanduser()
    if not path.is_file():
        raise FileNotFoundError(f"GnuCash book not found: {path}")
    sqlite_result = try_sqlite_book(path)
    if sqlite_result is not None:
        return sqlite_result
    return scan_xml_book(path)


def natural_key(value: str) -> list[Any]:
    parts = re.split(r"(\d+)", value)
    return [int(p) if p.isdigit() else p.lower() for p in parts]


def suggest_next_invoice_id(invoice_ids: Iterable[str], prefix: str = "", padding: int = 0) -> str:
    numbers: list[tuple[int, int]] = []
    if prefix:
        pattern = re.compile(r"^" + re.escape(prefix) + r"(\d+)$")
    else:
        pattern = re.compile(r"^(\d+)$")

    for invoice_id in invoice_ids:
        match = pattern.match(str(invoice_id).strip())
        if match:
            digits = match.group(1)
            numbers.append((int(digits), len(digits)))

    if not numbers:
        width = padding if padding > 0 else 0
        return prefix + (str(1).zfill(width) if width else "1")

    max_num = max(num for num, _width in numbers)
    detected_width = max(width for _num, width in numbers)
    width = padding if padding > 0 else detected_width
    return prefix + str(max_num + 1).zfill(width)


def increment_invoice_id(start: str, offset: int) -> str:
    match = re.match(r"^(.*?)(\d+)$", start)
    if not match:
        if offset == 0:
            return start
        return f"{start}-{offset + 1}"
    prefix, digits = match.groups()
    return f"{prefix}{str(int(digits) + offset).zfill(len(digits))}"


def customer_is_active(value: Any) -> bool:
    """Return a practical active/inactive flag for GnuCash customer rows.

    SQLite books normally use customers.active as 1 or 0. XML exports may use
    text values. Missing/blank active values are treated as active so older or
    unusual books do not hide customers unexpectedly.
    """
    raw = str(value or "").strip().lower()
    if raw in {"0", "false", "no", "n", "inactive"}:
        return False
    return True


def customers_to_json(customers: list[Customer]) -> list[dict[str, Any]]:
    return [
        {
            "id": c.id,
            "name": c.name,
            "active": c.active,
            "active_bool": customer_is_active(c.active),
            "guid": c.guid,
            "billing_name": c.billing_name,
        }
        for c in customers
    ]


def customer_balances_for_scan(book_path: str | Path, customers: list[Customer]) -> tuple[bool, dict[str, dict[str, str]]]:
    """Return ending customer balances from a SQLite book, keyed by customer ID.

    This is used by the web UI's "select non-zero balance" shortcut. It reuses
    the customer report data loader so the shortcut follows the same A/R logic
    used for batch statements. XML books can still be scanned for customers, but
    transaction-level balances currently require SQLite.
    """
    if not customers:
        return True, {}
    path = Path(book_path).expanduser()
    try:
        con = sqlite_connect_ro(path)
    except Exception:
        return False, {}
    try:
        customer_ids = [c.id for c in customers]
        data = load_report_data(con, customer_ids, date(1900, 1, 1), date(9999, 12, 31), set(), False)
    except Exception:
        return False, {}
    finally:
        con.close()

    balances: dict[str, dict[str, str]] = {}
    for customer in data.get("customers", []):
        total = Decimal("0")
        for account in customer.get("accounts", {}).values():
            for row in account.get("rows", []):
                total += row.get("amount", Decimal("0"))
        balances[str(customer.get("id") or "")] = {
            "balance": f"{total:.2f}",
            "nonzero_balance": "1" if total != 0 else "0",
        }
    # Customers without rows have a zero balance. Include them explicitly so the
    # UI knows balance scanning succeeded, even when no transactions exist.
    for customer in customers:
        balances.setdefault(customer.id, {"balance": "0.00", "nonzero_balance": "0"})
    return True, balances


def scan_book_json(args: argparse.Namespace) -> None:
    customers, invoice_ids = scan_book(args.book)
    accounts = scan_accounts(args.book)
    active_count = sum(1 for c in customers if customer_is_active(c.active))
    inactive_count = len(customers) - active_count
    customer_json = customers_to_json(customers)
    balance_scan_ok, balances = customer_balances_for_scan(args.book, customers)
    nonzero_count = 0
    for item in customer_json:
        balance = balances.get(str(item.get("id") or ""))
        if balance is not None:
            item.update(balance)
            if balance.get("nonzero_balance") == "1":
                nonzero_count += 1
    print_json(
        {
            "ok": True,
            "book": str(Path(args.book).expanduser()),
            "customer_count": len(customers),
            "active_customer_count": active_count,
            "inactive_customer_count": inactive_count,
            "customer_balance_scan_ok": balance_scan_ok,
            "nonzero_balance_customer_count": nonzero_count if balance_scan_ok else None,
            "invoice_count": len(invoice_ids),
            "next_invoice_id": suggest_next_invoice_id(invoice_ids, args.prefix or "", int(args.padding or 0)),
            "customers": customer_json,
            "account_count": len(accounts),
            "accounts": accounts,
            "income_accounts": [a for a in accounts if is_income_account(a)],
            "receivable_accounts": [a for a in accounts if is_receivable_account(a)],
        }
    )


def read_csv_rows(path: Path) -> list[list[str]]:
    data = path.read_text(errors="replace")
    sample = data[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [[cell.strip() for cell in row] for row in csv.reader(data.splitlines(), dialect)]


def read_xlsx_rows(path: Path) -> list[list[str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError("XLSX upload requires python3-openpyxl to be installed.") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if cell is None else str(cell).strip() for cell in row])
    return rows


def extract_ids_from_rows(rows: list[list[str]], known_ids: set[str]) -> list[str]:
    if not rows:
        return []
    # Header-driven mode.
    header = [cell.strip().lower() for cell in rows[0]]
    id_indexes = [idx for idx, name in enumerate(header) if name in CUSTOMER_HEADER_NAMES]
    found: list[str] = []
    if id_indexes:
        for row in rows[1:]:
            for idx in id_indexes:
                if idx < len(row):
                    value = row[idx].strip()
                    if value:
                        found.append(value)
        return dedupe(found)

    # Exact-match mode against known GnuCash customer IDs.
    for row in rows:
        for cell in row:
            cell = cell.strip()
            if cell in known_ids:
                found.append(cell)
    if found:
        return dedupe(found)

    # Fallback: first non-empty cell from each row.
    for row in rows:
        for cell in row:
            if cell.strip():
                found.append(cell.strip())
                break
    return dedupe(found)


def extract_ids_from_file(input_path: str | Path, known_ids: set[str]) -> list[str]:
    path = Path(input_path).expanduser()
    if not path.is_file():
        raise FileNotFoundError(f"Input file not found: {path}")
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        rows = read_xlsx_rows(path)
    elif suffix in {".csv", ".tsv"}:
        rows = read_csv_rows(path)
    else:
        text = path.read_text(errors="replace")
        rows = [[line.strip()] for line in text.splitlines() if line.strip()]
    return extract_ids_from_rows(rows, known_ids)


def dedupe(values: Iterable[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        clean = str(value).strip().strip('"').strip("'")
        if not clean or clean in seen:
            continue
        out.append(clean)
        seen.add(clean)
    return out


def match_customer_ids(ids: Iterable[str], customers: list[Customer]) -> dict[str, Any]:
    by_id = {c.id: c for c in customers}
    matched: list[dict[str, str]] = []
    unmatched: list[str] = []
    for cid in dedupe(ids):
        customer = by_id.get(cid)
        if customer is None:
            unmatched.append(cid)
        else:
            matched.append({"id": customer.id, "name": customer.name, "active": customer.active, "active_bool": customer_is_active(customer.active), "guid": customer.guid})
    return {"matched": matched, "unmatched": unmatched}


def scan_upload_json(args: argparse.Namespace) -> None:
    customers, invoice_ids = scan_book(args.book)
    known_ids = {c.id for c in customers}
    ids = extract_ids_from_file(args.input, known_ids)
    result = match_customer_ids(ids, customers)
    result.update(
        {
            "ok": True,
            "input_count": len(ids),
            "customer_count": len(customers),
            "invoice_count": len(invoice_ids),
            "next_invoice_id": suggest_next_invoice_id(invoice_ids, args.prefix or "", int(args.padding or 0)),
        }
    )
    print_json(result)


def scan_ids_json(args: argparse.Namespace) -> None:
    payload = json.load(sys.stdin)
    customers, invoice_ids = scan_book(args.book)
    ids = payload.get("customer_ids", [])
    result = match_customer_ids(ids, customers)
    result.update(
        {
            "ok": True,
            "input_count": len(ids),
            "customer_count": len(customers),
            "invoice_count": len(invoice_ids),
            "next_invoice_id": suggest_next_invoice_id(invoice_ids, args.prefix or "", int(args.padding or 0)),
        }
    )
    print_json(result)


def decimal_string(value: Any, default: str = "0") -> str:
    try:
        dec = Decimal(str(value).strip() or default)
    except (InvalidOperation, ValueError):
        dec = Decimal(default)
    # Use fixed-point plain string, removing unnecessary trailing zeroes.
    s = format(dec, "f")
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s or "0"


def money_string(value: Any, default: str = "0") -> str:
    """Return a money-style decimal string with exactly two fractional digits.

    GnuCash invoice import is sensitive to locale/decimal settings. Exporting
    prices as explicit fixed-point values, e.g. 15.00 instead of 15, avoids
    cases where whole-dollar prices are interpreted as cents.
    """
    try:
        dec = Decimal(str(value).strip() or default)
    except (InvalidOperation, ValueError):
        dec = Decimal(default)
    return format(dec.quantize(Decimal("0.01")), "f")



def quantity_string(value: Any, default: str = "1") -> str:
    """Return an explicit decimal quantity string with at least one digit.

    Some GnuCash import/date/number preference combinations can interpret an
    integer-looking quantity such as 1 as 0.01. Exporting 1.0 makes the decimal
    point explicit while still allowing fractional quantities such as 1.5.
    """
    try:
        dec = Decimal(str(value).strip() or default)
    except (InvalidOperation, ValueError):
        dec = Decimal(default)
    s = format(dec, "f")
    if "." not in s:
        return s + ".0"
    whole, frac = s.split(".", 1)
    frac = frac.rstrip("0")
    if frac == "":
        frac = "0"
    return whole + "." + frac

def yn(value: bool) -> str:
    return "Y" if bool(value) else "N"




def normalize_date_format(fmt: str) -> str:
    raw = str(fmt or "").strip()
    compact = raw.lower().replace(" ", "")
    mapping = {
        "m/d/y": "%m/%d/%Y",
        "mm/dd/yyyy": "%m/%d/%Y",
        "m/d/yyyy": "%m/%d/%Y",
        "%m/%d/%y": "%m/%d/%Y",
        "%m/%d/%Y".lower(): "%m/%d/%Y",
        "d/m/y": "%d/%m/%Y",
        "dd/mm/yyyy": "%d/%m/%Y",
        "d/m/yyyy": "%d/%m/%Y",
        "%d/%m/%y": "%d/%m/%Y",
        "%d/%m/%Y".lower(): "%d/%m/%Y",
        "y-m-d": "%Y-%m-%d",
        "yyyy-mm-dd": "%Y-%m-%d",
        "%Y-%m-%d".lower(): "%Y-%m-%d",
        "y/m/d": "%Y/%m/%d",
        "yyyy/mm/dd": "%Y/%m/%d",
        "d.m.y": "%d.%m.%Y",
        "dd.mm.yyyy": "%d.%m.%Y",
        "m-d-y": "%m-%d-%Y",
        "mm-dd-yyyy": "%m-%d-%Y",
        "d-m-y": "%d-%m-%Y",
        "dd-mm-yyyy": "%d-%m-%Y",
    }
    if compact in mapping:
        return mapping[compact]
    # Translate common PHP date tokens to strftime tokens.
    py = raw.replace("Y", "%Y").replace("y", "%y").replace("m", "%m").replace("n", "%m").replace("d", "%d").replace("j", "%d")
    if "%" in py:
        return py
    return "%m/%d/%Y"


def parse_form_date(value: Any) -> date | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(raw[:10] if fmt == "%Y-%m-%d" else raw, fmt).date()
        except ValueError:
            continue
    return None


def import_date(value: Any, fmt: str) -> str:
    dt = parse_form_date(value)
    if dt is None:
        return str(value or "")
    return dt.strftime(normalize_date_format(fmt))

def generate_json(args: argparse.Namespace) -> None:
    payload = json.load(sys.stdin)
    customers, invoice_ids = scan_book(args.book)
    known_ids = {c.id for c in customers}
    customer_ids = [cid for cid in dedupe(payload.get("customer_ids", [])) if cid in known_ids]
    params = payload.get("params", {}) or {}
    if not customer_ids:
        err("No valid customer IDs were supplied for generation.")

    start_id = str(params.get("start_invoice_id") or "").strip()
    if not start_id:
        start_id = suggest_next_invoice_id(invoice_ids, args.prefix or "", int(args.padding or 0))

    delimiter = str(params.get("csv_delimiter") or ",")
    if delimiter not in {",", ";"}:
        delimiter = ","

    posted = bool(params.get("posted", True))
    taxable = bool(params.get("taxable", False))
    taxincluded = bool(params.get("taxincluded", False))
    accu_splits = bool(params.get("accu_splits", False))
    csv_date_format = str(params.get("csv_date_format") or "m/d/Y")

    rows: list[list[str]] = []
    warnings: list[str] = []
    generated_ids: list[str] = []
    existing = set(invoice_ids)

    for idx, customer_id in enumerate(customer_ids):
        invoice_id = increment_invoice_id(start_id, idx)
        generated_ids.append(invoice_id)
        if invoice_id in existing:
            warnings.append(f"Generated invoice ID already exists in book: {invoice_id}")

        row = [
            invoice_id,
            import_date(params.get("date_opened") or "", csv_date_format),
            customer_id,
            str(params.get("billing_id") or ""),
            str(params.get("notes") or ""),
            import_date(params.get("entry_date") or params.get("date_opened") or "", csv_date_format),
            str(params.get("description") or ""),
            str(params.get("action") or "ea"),
            str(params.get("income_account") or ""),
            quantity_string(params.get("quantity", "1"), "1"),
            money_string(params.get("price", "0"), "0"),
            "%" if decimal_string(params.get("discount", "0"), "0") != "0" else "",
            "=" if decimal_string(params.get("discount", "0"), "0") != "0" else "",
            money_string(params.get("discount", "0"), "0"),
            yn(taxable),
            yn(taxincluded),
            str(params.get("tax_table") or ""),
            import_date(params.get("date_posted") or "", csv_date_format) if posted else "",
            import_date(params.get("due_date") or "", csv_date_format) if posted else "",
            str(params.get("ar_account") or "") if posted else "",
            str(params.get("memo_posted") or "") if posted else "",
            yn(accu_splits) if posted and accu_splits else "",
        ]
        rows.append(row)

    out = Path(args.out).expanduser()
    ensure_runtime_dir(out.parent)
    with out.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)
        writer.writerows(rows)
    apply_runtime_permissions(out)

    print_json(
        {
            "ok": True,
            "output": str(out),
            "invoice_count": len(customer_ids),
            "row_count": len(rows),
            "first_invoice_id": generated_ids[0] if generated_ids else "",
            "last_invoice_id": generated_ids[-1] if generated_ids else "",
            "warnings": warnings,
            "fields": FIELDNAMES,
        }
    )



# ---------------------------------------------------------------------------
# Customer report generation
# ---------------------------------------------------------------------------

REPORT_CSS = """
@page { size: __PAGE_SIZE__; margin: 0.45in; }
* { box-sizing: border-box; }
body { font-family: Arial, Helvetica, sans-serif; color: #222; font-size: 11pt; }
a { color: #1a3db7; text-decoration: underline; }
.report-header { display: grid; grid-template-columns: 1fr auto; gap: 1rem; align-items: start; margin-bottom: 1rem; }
.report-title { font-size: 18pt; font-weight: 700; margin: 0 0 .8rem 0; }
.report-meta { text-align: right; line-height: 1.8; }
.logo { max-height: 72px; max-width: 240px; display: block; margin-left: auto; margin-bottom: .4rem; }
.customer-name { margin: .2rem 0 .7rem 0; font-size: 12pt; }
.account-section { margin: 1.1rem 0 1.5rem 0; break-inside: auto; }
.account-title { font-size: 16pt; margin: 1rem 0 .6rem 0; break-after: avoid; }
table { border-collapse: collapse; width: 100%; }
thead { display: table-header-group; }
tr { break-inside: avoid; }
th, td { border: 1px solid #8a8a8a; padding: 4px 5px; vertical-align: top; }
th { font-weight: 700; text-align: center; background: #f4f4f4; }
td.num, th.num { text-align: right; white-space: nowrap; }
td.date, td.ref, td.type { white-space: nowrap; }
.period-total td { font-weight: 700; }
.balance-negative { color: #d00000; }
.money-link { color: #1a3db7; text-decoration: underline; font-style: italic; }
.total-row td { font-weight: 700; }
.aging-table { width: auto; margin: .55rem 0 1rem 0; break-inside: avoid; }
.aging-table th, .aging-table td { min-width: 4.8rem; }
.footer { margin-top: 1rem; color: #555; font-size: 9pt; }
@media print { body { padding-bottom: 0.2in; } }
.page-break { break-after: page; }
"""


def sqlite_connect_ro(book_path: Path) -> sqlite3.Connection:
    con = sqlite3.connect(f"file:{book_path}?mode=ro", uri=True)
    con.row_factory = sqlite3.Row
    return con


def table_columns(cur: sqlite3.Cursor, table: str) -> set[str]:
    return {row[1] for row in cur.execute(f"PRAGMA table_info({table})")}


def table_exists(cur: sqlite3.Cursor, table: str) -> bool:
    return cur.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone() is not None


def dec_from_num_denom(num: Any, denom: Any) -> Decimal:
    try:
        denom_d = Decimal(str(denom or 1))
        if denom_d == 0:
            denom_d = Decimal(1)
        return Decimal(str(num or 0)) / denom_d
    except Exception:
        return Decimal("0")


def money(value: Decimal | int | float | str) -> str:
    d = Decimal(str(value)).quantize(Decimal("0.01"))
    prefix = "-$" if d < 0 else "$"
    return f"{prefix}{abs(d):,.2f}"


def money_cell(value: Decimal) -> str:
    cls = "balance-negative" if value < 0 else ""
    return f'<span class="{cls}">{html_escape(money(value))}</span>'


def html_escape(value: Any) -> str:
    import html
    return html.escape(str(value or ""), quote=True)


def parse_iso_date(value: str, fallback: date | None = None) -> date:
    value = str(value or "").strip()[:10]
    try:
        return date.fromisoformat(value)
    except ValueError:
        return fallback or date.today()


def display_date(value: str) -> str:
    value = str(value or "").strip()[:10]
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%m/%d/%Y")
    except ValueError:
        return value


def account_full_names(con: sqlite3.Connection) -> dict[str, str]:
    cur = con.cursor()
    if not table_exists(cur, "accounts"):
        return {}
    rows = {row["guid"]: dict(row) for row in cur.execute("SELECT guid, name, parent_guid FROM accounts")}

    def full(guid: str, seen: set[str] | None = None) -> str:
        seen = seen or set()
        row = rows.get(guid)
        if not row:
            return guid
        name = str(row.get("name") or "")
        parent = str(row.get("parent_guid") or "")
        if not parent or parent in seen or parent not in rows:
            return name
        parent_name = full(parent, seen | {guid})
        if parent_name in {"Root Account", "Root", ""}:
            return name
        return parent_name + ":" + name

    return {guid: full(guid) for guid in rows}



def payment_counterpart_rows(con: sqlite3.Connection, tx_guid: str, ar_acct: str) -> list[sqlite3.Row]:
    """Return non-current-account splits for a transaction with account metadata.

    Customer reports should show real payments, not internal A/R lot allocation
    splits. A real payment has a counterpart in a cash/bank/asset/liability style
    account. Internal allocations usually have only RECEIVABLE counterpart splits.
    """
    cur = con.cursor()
    if not table_exists(cur, "splits") or not table_exists(cur, "accounts"):
        return []
    return list(cur.execute(
        """
        SELECT s.guid, s.memo, s.value_num, s.value_denom, s.account_guid,
               a.name AS account_name, a.account_type
        FROM splits s
        LEFT JOIN accounts a ON a.guid = s.account_guid
        WHERE s.tx_guid = ?
          AND s.account_guid != ?
        ORDER BY s.guid
        """,
        (tx_guid, ar_acct),
    ))


def is_actual_payment_transaction(con: sqlite3.Connection, tx_guid: str, ar_acct: str) -> bool:
    payment_types = {"BANK", "CASH", "ASSET", "CREDIT", "LIABILITY"}
    for row in payment_counterpart_rows(con, tx_guid, ar_acct):
        amount = dec_from_num_denom(row["value_num"], row["value_denom"])
        if amount == 0:
            continue
        account_type = str(row["account_type"] or "").upper()
        if account_type in payment_types:
            return True
    return False


def payment_counterparty_description(con: sqlite3.Connection, tx_guid: str, ar_acct: str, fallback: str, full_names: dict[str, str]) -> str:
    rows = payment_counterpart_rows(con, tx_guid, ar_acct)
    payment_types = {"BANK", "CASH", "ASSET", "CREDIT", "LIABILITY"}
    labels: list[str] = []
    for row in rows:
        account_type = str(row["account_type"] or "").upper()
        if account_type not in payment_types:
            continue
        memo = str(row["memo"] or "").strip()
        if memo and memo not in labels:
            labels.append(memo)
    if labels:
        return "\n".join(labels)
    if fallback:
        return fallback
    account_labels: list[str] = []
    for row in rows:
        account_type = str(row["account_type"] or "").upper()
        if account_type not in payment_types:
            continue
        name = full_names.get(str(row["account_guid"] or ""), str(row["account_name"] or "")).strip()
        if name and name not in account_labels:
            account_labels.append(name)
    return "\n".join(account_labels) if account_labels else "Payment"


def report_metadata_json(args: argparse.Namespace) -> None:
    path = Path(args.book).expanduser()
    try:
        con = sqlite_connect_ro(path)
    except sqlite3.Error as exc:
        raise RuntimeError(f"Report metadata requires a readable SQLite GnuCash book: {exc}") from exc
    try:
        cur = con.cursor()
        if not table_exists(cur, "accounts"):
            raise RuntimeError("No accounts table found. Report generation currently requires a SQLite GnuCash book.")
        accounts = scan_accounts(path)
        ar_accounts = [a for a in accounts if is_receivable_account(a)]
        income_accounts = [a for a in accounts if is_income_account(a)]
        print_json({"ok": True, "accounts": accounts, "ar_accounts": ar_accounts, "income_accounts": income_accounts})
    finally:
        con.close()


def load_report_data(con: sqlite3.Connection, customer_ids: list[str], date_from: date, date_to: date, selected_ar: set[str], include_internal_offsets: bool = False) -> dict[str, Any]:
    cur = con.cursor()
    needed = ["customers", "invoices", "accounts", "splits", "transactions"]
    missing = [t for t in needed if not table_exists(cur, t)]
    if missing:
        raise RuntimeError("Customer reports require a SQLite GnuCash book with these missing tables: " + ", ".join(missing))

    full_names = account_full_names(con)
    customers_by_id: dict[str, dict[str, Any]] = {}
    customers_by_guid: dict[str, dict[str, Any]] = {}
    customer_cols = table_columns(cur, "customers")
    name_expr = "name" if "name" in customer_cols else "id"
    billing_expr = "addr_name" if "addr_name" in customer_cols else "''"
    for row in cur.execute(f"SELECT guid, id, {name_expr} AS name, {billing_expr} AS billing_name FROM customers"):
        item = {
            "guid": row["guid"],
            "id": str(row["id"]),
            "name": str(row["name"] or row["id"]),
            "billing_name": str(row["billing_name"] or ""),
        }
        customers_by_id[item["id"]] = item
        customers_by_guid[item["guid"]] = item

    # Pull posted invoice/credit note A/R splits for selected customers.
    customer_guids = [customers_by_id[cid]["guid"] for cid in customer_ids if cid in customers_by_id]
    if not customer_guids:
        return {"customers": [], "accounts": []}

    qmarks = ",".join("?" for _ in customer_guids)
    invoice_rows = list(cur.execute(
        f"""
        SELECT i.guid, i.id, i.owner_guid, i.date_opened, i.date_posted, i.notes, i.billing_id,
               i.post_txn, i.post_lot, i.post_acc,
               t.post_date AS txn_post_date, t.description AS txn_description,
               s.value_num, s.value_denom, s.memo AS split_memo, s.account_guid
        FROM invoices i
        LEFT JOIN transactions t ON t.guid = i.post_txn
        LEFT JOIN splits s ON s.tx_guid = i.post_txn AND (s.account_guid = i.post_acc OR i.post_acc IS NULL OR i.post_acc = '')
        WHERE i.owner_guid IN ({qmarks})
          AND i.post_txn IS NOT NULL AND i.post_txn != ''
        """,
        customer_guids,
    ))

    by_customer: dict[str, dict[str, Any]] = {}
    invoice_txns: set[str] = set()
    lots_by_customer_account: dict[tuple[str, str], set[str]] = {}
    invoice_amounts: dict[str, Decimal] = {}

    for row in invoice_rows:
        cust = customers_by_guid.get(row["owner_guid"])
        if not cust:
            continue
        acct = str(row["account_guid"] or row["post_acc"] or "")
        if selected_ar and acct not in selected_ar:
            continue
        tx_date_raw = str(row["txn_post_date"] or row["date_posted"] or row["date_opened"] or "")[:10]
        tx_date = parse_iso_date(tx_date_raw)
        if tx_date < date_from or tx_date > date_to:
            continue
        amount = dec_from_num_denom(row["value_num"], row["value_denom"])
        inv_guid = str(row["guid"])
        if inv_guid in invoice_amounts:
            # Avoid duplicate A/R splits in odd books; keep first matching A/R row.
            continue
        invoice_amounts[inv_guid] = amount
        invoice_txns.add(str(row["post_txn"] or ""))
        lot = str(row["post_lot"] or "")
        if lot:
            lots_by_customer_account.setdefault((cust["id"], acct), set()).add(lot)
        desc = first_invoice_description(cur, inv_guid) or row["notes"] or row["billing_id"] or row["txn_description"] or "Invoice"
        row_item = {
            "date": tx_date_raw,
            "due_date": tx_date_raw,
            "reference": str(row["id"] or ""),
            "type": "Invoice" if amount >= 0 else "Credit Note",
            "description": str(desc),
            "debit": amount if amount > 0 else Decimal("0"),
            "credit": -amount if amount < 0 else Decimal("0"),
            "amount": amount,
            "sort": (tx_date_raw, 1, str(row["id"] or "")),
        }
        add_report_row(by_customer, cust, acct, full_names.get(acct, acct), row_item)

    # Payment/credit splits against lots belonging to each customer's posted invoices.
    # Default behavior intentionally suppresses internal A/R lot allocation/offset rows.
    # GnuCash often creates balancing splits when credits are applied to invoices; those
    # are useful for lot accounting but make customer statements noisy. We only show a
    # non-invoice split as a payment when the same transaction has a non-RECEIVABLE
    # payment-style counterpart such as Cash, Bank, Asset, Liability, or Credit.
    payment_buckets: dict[tuple[str, str, str], dict[str, Any]] = {}
    for (customer_id, acct), lots in lots_by_customer_account.items():
        if not lots:
            continue
        placeholders = ",".join("?" for _ in lots)
        params = [acct, *lots]
        for row in cur.execute(
            f"""
            SELECT s.guid, s.tx_guid, s.memo, s.value_num, s.value_denom, s.lot_guid,
                   t.post_date, t.num, t.description
            FROM splits s
            JOIN transactions t ON t.guid = s.tx_guid
            WHERE s.account_guid = ?
              AND s.lot_guid IN ({placeholders})
            ORDER BY t.post_date, t.num, s.guid
            """,
            params,
        ):
            tx_guid = str(row["tx_guid"] or "")
            if tx_guid in invoice_txns:
                continue
            if not include_internal_offsets and not is_actual_payment_transaction(con, tx_guid, acct):
                continue
            tx_date_raw = str(row["post_date"] or "")[:10]
            tx_date = parse_iso_date(tx_date_raw)
            if tx_date < date_from or tx_date > date_to:
                continue
            amount = dec_from_num_denom(row["value_num"], row["value_denom"])
            if amount == 0:
                continue
            cust = customers_by_id[customer_id]
            key = (customer_id, acct, tx_guid)
            fallback_desc = str(row["description"] or row["memo"] or "Payment")
            bucket = payment_buckets.setdefault(key, {
                "customer": cust,
                "account": acct,
                "account_name": full_names.get(acct, acct),
                "date": tx_date_raw,
                "due_date": "",
                "reference": str(row["num"] or ""),
                "description": payment_counterparty_description(con, tx_guid, acct, fallback_desc, full_names),
                "amount": Decimal("0"),
                "sort": (tx_date_raw, 2, str(row["num"] or ""), tx_guid),
                "internal_offset": not is_actual_payment_transaction(con, tx_guid, acct),
            })
            bucket["amount"] += amount

    for bucket in payment_buckets.values():
        amount = bucket["amount"]
        if amount == 0:
            continue
        row_type = "Payment" if amount < 0 else "Adjustment"
        if bucket.get("internal_offset"):
            row_type = "Internal Offset" if amount < 0 else "Internal Adjustment"
        row_item = {
            "date": bucket["date"],
            "due_date": bucket["due_date"],
            "reference": bucket["reference"],
            "type": row_type,
            "description": bucket["description"],
            "debit": amount if amount > 0 else Decimal("0"),
            "credit": -amount if amount < 0 else Decimal("0"),
            "amount": amount,
            "sort": bucket["sort"],
        }
        add_report_row(by_customer, bucket["customer"], bucket["account"], bucket["account_name"], row_item)

    return {"customers": list(by_customer.values())}


def first_invoice_description(cur: sqlite3.Cursor, invoice_guid: str) -> str:
    if not table_exists(cur, "entries"):
        return ""
    cols = table_columns(cur, "entries")
    if "invoice" not in cols:
        return ""
    desc_col = "description" if "description" in cols else "notes" if "notes" in cols else "''"
    row = cur.execute(f"SELECT {desc_col} AS description FROM entries WHERE invoice=? ORDER BY date LIMIT 1", (invoice_guid,)).fetchone()
    return str(row["description"] or "") if row else ""


def add_report_row(by_customer: dict[str, dict[str, Any]], cust: dict[str, Any], acct: str, acct_name: str, row_item: dict[str, Any]) -> None:
    c = by_customer.setdefault(cust["id"], {"id": cust["id"], "name": cust["name"], "billing_name": cust.get("billing_name", ""), "accounts": {}})
    a = c["accounts"].setdefault(acct, {"guid": acct, "name": acct_name, "rows": []})
    a["rows"].append(row_item)


def calc_aging(rows: list[dict[str, Any]], as_of: date) -> dict[str, Decimal]:
    # Conservative first pass: place the ending balance into an age bucket based on the oldest
    # unpaid positive invoice date. Negative balances are shown in Total only.
    balance = sum((r["amount"] for r in rows), Decimal("0"))
    buckets = {"prepayment": Decimal("0"), "current": Decimal("0"), "0-30": Decimal("0"), "31-60": Decimal("0"), "61-90": Decimal("0"), "91+": Decimal("0"), "total": balance}
    if balance < 0:
        buckets["prepayment"] = balance
        return buckets
    if balance == 0:
        return buckets
    invoice_dates = [parse_iso_date(r.get("due_date") or r.get("date") or "") for r in rows if r.get("debit", Decimal("0")) > 0]
    age = (as_of - min(invoice_dates or [as_of])).days
    if age <= 0:
        key = "current"
    elif age <= 30:
        key = "0-30"
    elif age <= 60:
        key = "31-60"
    elif age <= 90:
        key = "61-90"
    else:
        key = "91+"
    buckets[key] = balance
    return buckets


def render_customer_html(customer: dict[str, Any], payload: dict[str, Any], style_css: str) -> str:
    as_of = parse_iso_date(payload.get("date_to", ""))
    page_size = str(payload.get("page_size") or "Letter")
    css = REPORT_CSS.replace("__PAGE_SIZE__", page_size) + "\n" + style_css + "\n" + str(payload.get("custom_css") or "")
    logo = ""
    logo_path = str(payload.get("logo_path") or "")
    if logo_path and Path(logo_path).is_file():
        logo_uri = Path(logo_path).resolve().as_uri()
        logo = f'<img class="logo" src="{html_escape(logo_uri)}" alt="Logo">'
    header = f"""
    <div class="report-header">
      <div>
        <h1 class="report-title">Customer Report: <a>{html_escape(customer['name'])}</a></h1>
        <div class="customer-name">{html_escape(customer['name'])}</div>
        <div>Date Range: {html_escape(display_date(payload.get('date_from','')))} - {html_escape(display_date(payload.get('date_to','')))}</div>
      </div>
      <div class="report-meta">{logo}<div>{html_escape(payload.get('organization_name',''))}</div><div>{html_escape(display_date(payload.get('date_to','')))}</div></div>
    </div>
    """
    sections = []
    for account in sorted(customer.get("accounts", {}).values(), key=lambda a: a.get("name", "")):
        rows = sorted(account["rows"], key=lambda r: r["sort"])
        if not rows:
            continue
        balance = Decimal("0")
        debit_total = Decimal("0")
        credit_total = Decimal("0")
        body_rows = []
        for r in rows:
            debit_total += r["debit"]
            credit_total += r["credit"]
            balance += r["amount"]
            body_rows.append(
                "<tr>"
                f"<td class=\"date\">{html_escape(display_date(r['date']))}</td>"
                f"<td class=\"date\">{html_escape(display_date(r.get('due_date','')))}</td>"
                f"<td class=\"ref\"><a>{html_escape(r.get('reference',''))}</a></td>"
                f"<td class=\"type\">{html_escape(r.get('type',''))}</td>"
                f"<td>{html_escape(r.get('description',''))}</td>"
                f"<td class=\"num\">{('<span class=\"money-link\">' + html_escape(money(r['debit'])) + '</span>') if r['debit'] else ''}</td>"
                f"<td class=\"num\">{('<span class=\"money-link\">' + html_escape(money(r['credit'])) + '</span>') if r['credit'] else ''}</td>"
                f"<td class=\"num\">{money_cell(balance)}</td>"
                "</tr>"
            )
        aging = calc_aging(rows, as_of)
        due_label = "Total Credit" if balance < 0 else "Total Due"
        sections.append(
            f"<section class=\"account-section\"><h2 class=\"account-title\">Account: {html_escape(account['name'])}</h2>"
            "<table><thead><tr><th>Date</th><th>Due Date</th><th>Reference</th><th>Type</th><th>Description</th><th class=\"num\">Debits</th><th class=\"num\">Credits</th><th class=\"num\">Balance</th></tr></thead><tbody>"
            + "\n".join(body_rows)
            + f"<tr class=\"period-total\"><td colspan=\"5\">Period<br>Totals</td><td class=\"num\">{html_escape(money(debit_total))}</td><td class=\"num\">{html_escape(money(credit_total))}</td><td class=\"num\">{money_cell(balance)}</td></tr>"
            + f"<tr class=\"total-row\"><td colspan=\"7\">{due_label}</td><td class=\"num\">{money_cell(balance)}</td></tr></tbody></table>"
            + "<table class=\"aging-table\"><thead><tr><th>Pre-Payment</th><th>Current</th><th>0-30 days</th><th>31-60 days</th><th>61-90 days</th><th>91+ days</th><th>Total</th></tr></thead><tbody><tr>"
            + f"<td>{html_escape(money(aging['prepayment']))}</td><td>{html_escape(money(aging['current']))}</td><td>{html_escape(money(aging['0-30']))}</td><td>{html_escape(money(aging['31-60']))}</td><td>{html_escape(money(aging['61-90']))}</td><td>{html_escape(money(aging['91+']))}</td><td>{html_escape(money(aging['total']))}</td>"
            + "</tr></tbody></table></section>"
        )
    footer = f"<div class=\"footer\">{html_escape(payload.get('footer_text',''))}</div>" if payload.get("footer_text") else ""
    return "<!doctype html><html><head><meta charset=\"utf-8\"><style>" + css + "</style></head><body>" + header + "\n".join(sections) + footer + "</body></html>"


def extract_style_reference(path: str) -> str:
    if not path:
        return ""
    p = Path(path)
    if not p.is_file():
        return ""
    text = p.read_text(errors="replace")
    styles = re.findall(r"<style[^>]*>(.*?)</style>", text, flags=re.I | re.S)
    return "\n".join(styles)


def safe_filename(value: str) -> str:
    clean = str(value or "").strip()
    clean = re.sub(r'[<>:"/\\|?*\x00-\x1F]+', "-", clean)
    clean = re.sub(r"\s+", " ", clean)
    clean = re.sub(r"-{3,}", "--", clean)
    clean = clean.strip(" .-_")
    if not clean:
        clean = "customer"
    return clean[:180]


def filename_customer_value(customer: dict[str, Any], source: str) -> str:
    source = str(source or "billing_name").strip().lower()
    if source in {"customer_id", "customer_number", "id", "number"}:
        return str(customer.get("id") or "")
    if source in {"company_name", "company", "name"}:
        return str(customer.get("name") or customer.get("id") or "")
    if source in {"billing_name", "billing", "addr_name"}:
        return str(customer.get("billing_name") or customer.get("name") or customer.get("id") or "")
    return str(customer.get("billing_name") or customer.get("name") or customer.get("id") or "")


def filename_date(value: Any, fmt: str) -> str:
    dt = parse_form_date(value)
    if dt is None:
        return str(value or "")
    return dt.strftime(normalize_date_format(fmt or "Y-m-d"))


def render_report_filename(customer: dict[str, Any], payload: dict[str, Any]) -> str:
    template = str(payload.get("filename_template") or "{customer} - {date_to} - {text}")
    source = str(payload.get("filename_customer_source") or "billing_name")
    date_fmt = str(payload.get("filename_date_format") or "Y-m-d")
    date_from = filename_date(payload.get("date_from", ""), date_fmt)
    date_to = filename_date(payload.get("date_to", ""), date_fmt)
    selected_customer = filename_customer_value(customer, source)
    values = {
        "customer": selected_customer,
        "customer_id": str(customer.get("id") or ""),
        "customer_number": str(customer.get("id") or ""),
        "company_name": str(customer.get("name") or ""),
        "name": str(customer.get("name") or ""),
        "billing_name": str(customer.get("billing_name") or ""),
        "date": date_to,
        "date_to": date_to,
        "date_from": date_from,
        "group": str(payload.get("group_name") or ""),
        "text": str(payload.get("filename_text") or "statement"),
    }

    def repl(match: re.Match[str]) -> str:
        key = match.group(1).strip().lower()
        return values.get(key, "")

    rendered = re.sub(r"\{([A-Za-z0-9_ -]+)\}", repl, template)
    rendered = re.sub(r"\s+", " ", rendered).strip()
    return safe_filename(rendered)


def chromium_candidates(configured: str = "") -> list[str]:
    candidates: list[str] = []

    def add(value: str | None) -> None:
        value = str(value or "").strip()
        if value and value not in candidates:
            candidates.append(value)

    add(configured)
    # Ubuntu's chromium package often launches the snap; the real command is
    # commonly /snap/bin/chromium on local desktops. Try it before the older
    # /usr/bin paths when the configured value is missing or stale.
    add("/snap/bin/chromium")
    add("/usr/bin/chromium")
    add("/usr/bin/chromium-browser")
    add("/usr/bin/google-chrome")
    add("/usr/bin/google-chrome-stable")
    add("/opt/google/chrome/chrome")
    for name in ["chromium", "chromium-browser", "google-chrome", "google-chrome-stable"]:
        add(shutil.which(name))
    return candidates


def resolve_chromium_binary(configured: str = "") -> str:
    tried = chromium_candidates(configured)
    for candidate in tried:
        path = Path(candidate)
        if path.is_file() and os.access(path, os.X_OK):
            return str(path)
        # If the user entered a bare command name, resolve it through PATH.
        if os.sep not in candidate:
            found = shutil.which(candidate)
            if found and Path(found).is_file() and os.access(found, os.X_OK):
                return found
    raise RuntimeError(
        "Chromium binary not found. Install chromium or set Chromium binary in Settings. "
        "Tried: " + ", ".join(tried)
    )



def _free_tcp_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind(("127.0.0.1", 0))
        return int(sock.getsockname()[1])


def _http_json(url: str, *, method: str = "GET", timeout: float = 1.0) -> Any:
    req = urllib.request.Request(url, method=method)
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def _wait_for_devtools(port: int, timeout: float = 12.0) -> None:
    deadline = time.monotonic() + timeout
    last_error = ""
    while time.monotonic() < deadline:
        try:
            _http_json(f"http://127.0.0.1:{port}/json/version", timeout=0.75)
            return
        except Exception as exc:  # pragma: no cover - depends on Chromium startup timing
            last_error = str(exc)
            time.sleep(0.15)
    raise RuntimeError(f"Chromium DevTools endpoint did not become ready: {last_error}")


class _WebSocketCDP:
    """Tiny standard-library WebSocket client for Chromium DevTools JSON-RPC.

    This avoids adding a Python package dependency just to use Chromium's
    PrintToPDF header/footer templates. Client-to-server frames must be masked;
    Chromium's server-to-client frames are not masked.
    """

    def __init__(self, websocket_url: str):
        parsed = urllib.parse.urlparse(websocket_url)
        if parsed.scheme not in {"ws", "wss"}:
            raise RuntimeError(f"Unsupported DevTools websocket URL: {websocket_url}")
        if parsed.scheme == "wss":
            raise RuntimeError("wss DevTools endpoints are not supported by the built-in client")
        self.host = parsed.hostname or "127.0.0.1"
        self.port = int(parsed.port or 80)
        self.path = parsed.path or "/"
        if parsed.query:
            self.path += "?" + parsed.query
        self.sock = socket.create_connection((self.host, self.port), timeout=10)
        self.sock.settimeout(30)
        self._next_id = 1
        self._handshake()

    def _handshake(self) -> None:
        key = base64.b64encode(secrets.token_bytes(16)).decode("ascii")
        request = (
            f"GET {self.path} HTTP/1.1\r\n"
            f"Host: {self.host}:{self.port}\r\n"
            "Upgrade: websocket\r\n"
            "Connection: Upgrade\r\n"
            f"Sec-WebSocket-Key: {key}\r\n"
            "Sec-WebSocket-Version: 13\r\n\r\n"
        )
        self.sock.sendall(request.encode("ascii"))
        response = b""
        while b"\r\n\r\n" not in response:
            chunk = self.sock.recv(4096)
            if not chunk:
                break
            response += chunk
        if b" 101 " not in response.split(b"\r\n", 1)[0]:
            raise RuntimeError("Chromium DevTools websocket handshake failed")
        accept_expected = base64.b64encode(
            hashlib.sha1((key + "258EAFA5-E914-47DA-95CA-C5AB0DC85B11").encode("ascii")).digest()
        ).decode("ascii")
        if accept_expected.encode("ascii") not in response:
            raise RuntimeError("Chromium DevTools websocket handshake returned an unexpected accept key")

    def close(self) -> None:
        try:
            self.sock.close()
        except OSError:
            pass

    def _read_exact(self, n: int) -> bytes:
        chunks: list[bytes] = []
        remaining = n
        while remaining > 0:
            chunk = self.sock.recv(remaining)
            if not chunk:
                raise RuntimeError("Chromium DevTools websocket closed unexpectedly")
            chunks.append(chunk)
            remaining -= len(chunk)
        return b"".join(chunks)

    def _recv_frame(self) -> dict[str, Any]:
        while True:
            b1, b2 = self._read_exact(2)
            opcode = b1 & 0x0F
            length = b2 & 0x7F
            masked = bool(b2 & 0x80)
            if length == 126:
                length = struct.unpack("!H", self._read_exact(2))[0]
            elif length == 127:
                length = struct.unpack("!Q", self._read_exact(8))[0]
            mask = self._read_exact(4) if masked else b""
            payload = self._read_exact(length) if length else b""
            if masked:
                payload = bytes(b ^ mask[i % 4] for i, b in enumerate(payload))
            if opcode == 8:
                raise RuntimeError("Chromium DevTools websocket closed")
            if opcode == 9:  # ping; reply with pong
                self._send_frame(payload, opcode=10)
                continue
            if opcode == 10:
                continue
            if opcode != 1:
                continue
            return json.loads(payload.decode("utf-8"))

    def _send_frame(self, payload: bytes, *, opcode: int = 1) -> None:
        header = bytearray([0x80 | opcode])
        length = len(payload)
        if length < 126:
            header.append(0x80 | length)
        elif length < (1 << 16):
            header.append(0x80 | 126)
            header.extend(struct.pack("!H", length))
        else:
            header.append(0x80 | 127)
            header.extend(struct.pack("!Q", length))
        mask = secrets.token_bytes(4)
        masked = bytes(b ^ mask[i % 4] for i, b in enumerate(payload))
        self.sock.sendall(bytes(header) + mask + masked)

    def send(self, method: str, params: dict[str, Any] | None = None) -> int:
        msg_id = self._next_id
        self._next_id += 1
        payload = {"id": msg_id, "method": method}
        if params is not None:
            payload["params"] = params
        self._send_frame(json.dumps(payload, separators=(",", ":")).encode("utf-8"))
        return msg_id

    def wait_response(self, msg_id: int, timeout: float = 30.0) -> dict[str, Any]:
        deadline = time.monotonic() + timeout
        while time.monotonic() < deadline:
            msg = self._recv_frame()
            if msg.get("id") == msg_id:
                if "error" in msg:
                    raise RuntimeError("Chromium DevTools error: " + json.dumps(msg["error"]))
                return msg.get("result", {})
        raise RuntimeError(f"Timed out waiting for Chromium DevTools response {msg_id}")

    def wait_event(self, method: str, timeout: float = 30.0) -> dict[str, Any]:
        deadline = time.monotonic() + timeout
        while time.monotonic() < deadline:
            msg = self._recv_frame()
            if msg.get("method") == method:
                return msg.get("params", {})
        raise RuntimeError(f"Timed out waiting for Chromium DevTools event {method}")


def _render_pdf_with_devtools(binary: str, html_path: Path, pdf_path: Path, *, show_page_numbers: bool) -> None:
    port = _free_tcp_port()
    html_uri = str(html_path.resolve().as_uri())
    with tempfile.TemporaryDirectory(prefix="gnc-batch-chromium-") as user_data_dir:
        cmd = [
            binary,
            "--headless=new",
            "--disable-gpu",
            "--no-sandbox",
            "--disable-dev-shm-usage",
            f"--remote-debugging-port={port}",
            "--remote-debugging-address=127.0.0.1",
            f"--user-data-dir={user_data_dir}",
            "about:blank",
        ]
        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        ws: _WebSocketCDP | None = None
        try:
            try:
                _wait_for_devtools(port)
            except Exception:
                # Some distro Chromium builds do not support --headless=new.
                proc.terminate()
                try:
                    proc.wait(timeout=4)
                except subprocess.TimeoutExpired:
                    proc.kill()
                cmd[1] = "--headless"
                proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
                _wait_for_devtools(port)

            target = _http_json(f"http://127.0.0.1:{port}/json/new?{urllib.parse.quote(html_uri, safe=':/')}", method="PUT", timeout=4)
            ws_url = target.get("webSocketDebuggerUrl")
            if not ws_url:
                raise RuntimeError("Chromium did not return a page DevTools websocket URL")
            ws = _WebSocketCDP(str(ws_url))
            ws.wait_response(ws.send("Page.enable"), timeout=10)
            ws.wait_response(ws.send("Runtime.enable"), timeout=10)
            # The /json/new URL normally loads the target, but navigate explicitly
            # to avoid races on slower snap Chromium launches.
            ws.send("Page.navigate", {"url": html_uri})
            ws.wait_event("Page.loadEventFired", timeout=30)
            # Give image assets and fonts a brief chance to settle before printing.
            time.sleep(0.25)
            params: dict[str, Any] = {
                "printBackground": True,
                "preferCSSPageSize": True,
                "displayHeaderFooter": bool(show_page_numbers),
            }
            if show_page_numbers:
                params["headerTemplate"] = "<span></span>"
                params["footerTemplate"] = (
                    "<div style='font-size:8pt;color:#555;width:100%;"
                    "padding:0 0.45in;text-align:right;'>"
                    "Page <span class='pageNumber'></span> of <span class='totalPages'></span>"
                    "</div>"
                )
            result = ws.wait_response(ws.send("Page.printToPDF", params), timeout=45)
            data = result.get("data")
            if not data:
                raise RuntimeError("Chromium returned no PDF data")
            pdf_path.write_bytes(base64.b64decode(data))
        finally:
            if ws is not None:
                ws.close()
            if proc.poll() is None:
                proc.terminate()
                try:
                    proc.wait(timeout=5)
                except subprocess.TimeoutExpired:
                    proc.kill()
        if not pdf_path.is_file() or pdf_path.stat().st_size == 0:
            raise RuntimeError("Chromium DevTools PDF render produced no output")


def _render_pdf_with_cli(binary: str, html_path: Path, pdf_path: Path) -> None:
    cmd = [
        binary,
        "--headless=new",
        "--disable-gpu",
        "--no-sandbox",
        "--disable-dev-shm-usage",
        "--no-pdf-header-footer",
        "--print-to-pdf-no-header",
        f"--print-to-pdf={pdf_path}",
        str(html_path.resolve().as_uri()),
    ]
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=120)
    if proc.returncode != 0 or not pdf_path.is_file():
        # Some distro Chromium builds do not support --headless=new.
        cmd[1] = "--headless"
        proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=120)
    if proc.returncode != 0 or not pdf_path.is_file():
        raise RuntimeError("Chromium PDF render failed using " + binary + ": " + (proc.stderr or proc.stdout or "unknown error"))


def render_pdf(chromium_bin: str, html_path: Path, pdf_path: Path, *, show_page_numbers: bool = True) -> None:
    binary = resolve_chromium_binary(chromium_bin)
    if show_page_numbers:
        try:
            _render_pdf_with_devtools(binary, html_path, pdf_path, show_page_numbers=True)
            return
        except Exception as exc:
            # The CLI fallback still suppresses Chromium's path/date headers, but
            # cannot render custom Page X of Y footers. Failing hard would block
            # report generation on systems where remote debugging is unavailable.
            sys.stderr.write(f"Warning: Chromium DevTools pagination failed; falling back without page numbers: {exc}\n")
    _render_pdf_with_cli(binary, html_path, pdf_path)

def customer_reports_json(args: argparse.Namespace) -> None:
    payload = json.load(sys.stdin)
    book = Path(args.book).expanduser()
    out_dir = Path(args.out_dir).expanduser()
    html_dir = out_dir / "html"
    pdf_dir = out_dir / "pdf"
    ensure_runtime_dir(html_dir)
    ensure_runtime_dir(pdf_dir)
    customer_ids = dedupe(payload.get("customer_ids", []))
    date_from = parse_iso_date(str(payload.get("date_from") or date.today().replace(month=1, day=1)))
    date_to = parse_iso_date(str(payload.get("date_to") or date.today()))
    selected_ar = set(map(str, payload.get("ar_accounts") or []))
    style_css = extract_style_reference(str(payload.get("style_reference_path") or ""))

    try:
        con = sqlite_connect_ro(book)
    except sqlite3.Error as exc:
        raise RuntimeError(f"Customer report generation requires a readable SQLite GnuCash book: {exc}") from exc
    try:
        data = load_report_data(con, customer_ids, date_from, date_to, selected_ar, bool(payload.get("show_internal_offsets", False)))
    finally:
        con.close()

    include_zero = bool(payload.get("include_zero_balance", True))
    generated: list[dict[str, Any]] = []
    used_names: set[str] = set()
    for customer in data["customers"]:
        # Skip customers with no account sections unless requested. At this stage no-row customers are not present.
        html = render_customer_html(customer, payload, style_css)
        base = render_report_filename(customer, payload)
        unique_base = base
        n = 2
        while unique_base.lower() in used_names:
            unique_base = safe_filename(f"{base} ({n})")
            n += 1
        used_names.add(unique_base.lower())
        html_path = html_dir / f"{unique_base}.html"
        pdf_path = pdf_dir / f"{unique_base}.pdf"
        html_path.write_text(html, encoding="utf-8")
        apply_runtime_permissions(html_path)
        if not include_zero:
            total = Decimal("0")
            for account in customer.get("accounts", {}).values():
                for row in account.get("rows", []):
                    total += row.get("amount", Decimal("0"))
            if total == 0:
                continue
        render_pdf(str(payload.get("chromium_bin") or ""), html_path, pdf_path, show_page_numbers=bool(payload.get("show_page_numbers", True)))
        apply_runtime_permissions(pdf_path)
        generated.append({"customer_id": customer["id"], "customer_name": customer["name"], "billing_name": customer.get("billing_name", ""), "html": str(html_path), "pdf": str(pdf_path), "status": "generated"})

    zip_path = out_dir / "customer-reports.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for item in generated:
            zf.write(item["pdf"], arcname=Path(item["pdf"]).name)
    apply_runtime_permissions(zip_path)

    manifest = {
        "ok": True,
        "book": str(book),
        "group_name": payload.get("group_name", ""),
        "date_from": str(payload.get("date_from", "")),
        "date_to": str(payload.get("date_to", "")),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "pdf_count": len(generated),
        "zip": str(zip_path),
        "customers": generated,
    }
    manifest_path = out_dir / "batch-manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    apply_runtime_permissions(manifest_path)
    apply_runtime_permissions(out_dir)
    print_json(manifest)


# ---------------------------------------------------------------------------
# Per-invoice PDF export
# ---------------------------------------------------------------------------

INVOICE_CSS = """
@page { size: __PAGE_SIZE__; margin: 0.45in; }
* { box-sizing: border-box; }
body { font-family: Arial, Helvetica, sans-serif; color: #222; font-size: 11pt; }
.invoice-header { display: grid; grid-template-columns: 1fr auto; gap: 1rem; align-items: start; margin-bottom: 1.1rem; }
.invoice-title { font-size: 22pt; font-weight: 700; margin: 0 0 .5rem 0; }
.invoice-meta { text-align: right; line-height: 1.7; }
.logo { max-height: 72px; max-width: 240px; display: block; margin-left: auto; margin-bottom: .4rem; }
.bill-to { border: 1px solid #888; padding: .65rem; margin: .8rem 0 1rem 0; width: 48%; min-height: 5rem; }
.bill-to h2 { font-size: 11pt; margin: 0 0 .35rem 0; color: #555; }
table { border-collapse: collapse; width: 100%; margin-top: .6rem; }
thead { display: table-header-group; }
tr { break-inside: avoid; }
th, td { border: 1px solid #8a8a8a; padding: 5px 6px; vertical-align: top; }
th { font-weight: 700; text-align: center; background: #f4f4f4; }
td.num, th.num { text-align: right; white-space: nowrap; }
td.date { white-space: nowrap; }
.invoice-total td { font-weight: 700; font-size: 12pt; }
.notes { margin-top: 1rem; white-space: pre-wrap; }
.footer { margin-top: 1rem; color: #555; font-size: 9pt; }
"""


def _first_existing(cols: set[str], names: list[str], default: str = "") -> str:
    for name in names:
        if name in cols:
            return name
    return default


def _entry_amount(entry: sqlite3.Row, cols: set[str]) -> Decimal:
    qn = entry["quantity_num"] if "quantity_num" in cols else 1
    qd = entry["quantity_denom"] if "quantity_denom" in cols else 1
    quantity = dec_from_num_denom(qn, qd)
    price_num_col = _first_existing(cols, ["i_price_num", "price_num", "b_price_num"])
    price_den_col = _first_existing(cols, ["i_price_denom", "price_denom", "b_price_denom"])
    price = dec_from_num_denom(entry[price_num_col], entry[price_den_col]) if price_num_col and price_den_col else Decimal("0")
    discount_num_col = _first_existing(cols, ["i_discount_num", "discount_num", "b_discount_num"])
    discount_den_col = _first_existing(cols, ["i_discount_denom", "discount_denom", "b_discount_denom"])
    discount = dec_from_num_denom(entry[discount_num_col], entry[discount_den_col]) if discount_num_col and discount_den_col else Decimal("0")
    return (quantity * price) - discount


def _invoice_posted_total(cur: sqlite3.Cursor, invoice: sqlite3.Row) -> Decimal | None:
    post_txn = str(invoice["post_txn"] or "") if "post_txn" in invoice.keys() else ""
    post_acc = str(invoice["post_acc"] or "") if "post_acc" in invoice.keys() else ""
    if not post_txn or not table_exists(cur, "splits"):
        return None
    cols = table_columns(cur, "splits")
    if not {"tx_guid", "value_num", "value_denom"}.issubset(cols):
        return None
    if post_acc and "account_guid" in cols:
        row = cur.execute(
            "SELECT value_num, value_denom FROM splits WHERE tx_guid=? AND account_guid=? LIMIT 1",
            (post_txn, post_acc),
        ).fetchone()
    else:
        row = cur.execute("SELECT value_num, value_denom FROM splits WHERE tx_guid=? LIMIT 1", (post_txn,)).fetchone()
    if not row:
        return None
    return dec_from_num_denom(row["value_num"], row["value_denom"])


def load_invoice_export_data(
    con: sqlite3.Connection,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
    customer_ids: list[str] | None = None,
    invoice_guids: list[str] | None = None,
) -> list[dict[str, Any]]:
    cur = con.cursor()
    needed = ["customers", "invoices"]
    missing = [t for t in needed if not table_exists(cur, t)]
    if missing:
        raise RuntimeError("Invoice PDF export requires a SQLite GnuCash book with these missing tables: " + ", ".join(missing))

    inv_cols = table_columns(cur, "invoices")
    cust_cols = table_columns(cur, "customers")
    entry_cols = table_columns(cur, "entries") if table_exists(cur, "entries") else set()
    full_names = account_full_names(con)

    cust_name_col = _first_existing(cust_cols, ["name", "id"], "id")
    cust_billing_col = _first_existing(cust_cols, ["addr_name", "name", "id"], cust_name_col)
    customers_by_guid: dict[str, dict[str, Any]] = {}
    for row in cur.execute(f"SELECT guid, id, {cust_name_col} AS name, {cust_billing_col} AS billing_name FROM customers"):
        customers_by_guid[str(row["guid"])] = {
            "guid": str(row["guid"]),
            "id": str(row["id"]),
            "name": str(row["name"] or row["id"]),
            "billing_name": str(row["billing_name"] or row["name"] or row["id"]),
        }

    where: list[str] = []
    params: list[Any] = []
    if invoice_guids:
        where.append("i.guid IN (" + ",".join("?" for _ in invoice_guids) + ")")
        params.extend(invoice_guids)
    if customer_ids:
        customer_guid_filter = [guid for guid, cust in customers_by_guid.items() if cust.get("id") in set(customer_ids)]
        if not customer_guid_filter:
            return []
        where.append("i.owner_guid IN (" + ",".join("?" for _ in customer_guid_filter) + ")")
        params.extend(customer_guid_filter)
    if date_from is not None:
        where.append("substr(COALESCE(NULLIF(i.date_posted,''), NULLIF(i.date_opened,'')), 1, 10) >= ?")
        params.append(date_from.isoformat())
    if date_to is not None:
        where.append("substr(COALESCE(NULLIF(i.date_posted,''), NULLIF(i.date_opened,'')), 1, 10) <= ?")
        params.append(date_to.isoformat())
    where_sql = " WHERE " + " AND ".join(where) if where else ""

    rows = list(cur.execute(
        "SELECT i.* FROM invoices i" + where_sql + " ORDER BY i.owner_guid, COALESCE(NULLIF(i.date_posted,''), NULLIF(i.date_opened,'')), i.id",
        params,
    ))
    invoices: list[dict[str, Any]] = []
    for inv in rows:
        owner_guid = str(inv["owner_guid"] or "") if "owner_guid" in inv.keys() else ""
        customer = customers_by_guid.get(owner_guid)
        if not customer:
            continue
        inv_guid = str(inv["guid"])
        inv_date = str((inv["date_posted"] if "date_posted" in inv.keys() else "") or (inv["date_opened"] if "date_opened" in inv.keys() else ""))[:10]
        due_date = ""
        for col in ["date_due", "due_date"]:
            if col in inv.keys() and inv[col]:
                due_date = str(inv[col])[:10]
                break
        if not due_date:
            due_date = inv_date
        entries: list[dict[str, Any]] = []
        if entry_cols and "invoice" in entry_cols:
            desc_col = _first_existing(entry_cols, ["description", "notes"], "''")
            action_col = _first_existing(entry_cols, ["action"], "''")
            acct_col = _first_existing(entry_cols, ["i_acct", "account", "b_acct"], "''")
            for entry in cur.execute(f"SELECT * FROM entries WHERE invoice=? ORDER BY date, guid", (inv_guid,)):
                quantity = dec_from_num_denom(entry["quantity_num"] if "quantity_num" in entry_cols else 1, entry["quantity_denom"] if "quantity_denom" in entry_cols else 1)
                price_num_col = _first_existing(entry_cols, ["i_price_num", "price_num", "b_price_num"])
                price_den_col = _first_existing(entry_cols, ["i_price_denom", "price_denom", "b_price_denom"])
                price = dec_from_num_denom(entry[price_num_col], entry[price_den_col]) if price_num_col and price_den_col else Decimal("0")
                amount = _entry_amount(entry, entry_cols)
                acct_guid = str(entry[acct_col] or "") if acct_col and acct_col != "''" else ""
                entries.append({
                    "date": str(entry["date"] if "date" in entry_cols else inv_date)[:10],
                    "description": str(entry[desc_col] or "") if desc_col != "''" else "",
                    "action": str(entry[action_col] or "") if action_col != "''" else "",
                    "account": full_names.get(acct_guid, acct_guid),
                    "quantity": quantity,
                    "unit_price": price,
                    "amount": amount,
                })
        entry_total = sum((e["amount"] for e in entries), Decimal("0"))
        posted_total = _invoice_posted_total(cur, inv)
        total = posted_total if posted_total is not None else entry_total
        description = first_invoice_description(cur, inv_guid) if table_exists(cur, "entries") else ""
        if not description:
            description = str((inv["notes"] if "notes" in inv.keys() else "") or (inv["billing_id"] if "billing_id" in inv.keys() else "") or "Invoice")
        invoices.append({
            "guid": inv_guid,
            "id": str(inv["id"] or ""),
            "customer": customer,
            "customer_id": customer["id"],
            "customer_name": customer["name"],
            "billing_name": customer.get("billing_name", ""),
            "date": inv_date,
            "due_date": due_date,
            "notes": str(inv["notes"] if "notes" in inv.keys() else ""),
            "billing_id": str(inv["billing_id"] if "billing_id" in inv.keys() else ""),
            "post_acc": str(inv["post_acc"] if "post_acc" in inv.keys() else ""),
            "ar_account": full_names.get(str(inv["post_acc"] if "post_acc" in inv.keys() else ""), str(inv["post_acc"] if "post_acc" in inv.keys() else "")),
            "description": description,
            "entries": entries,
            "total": total,
            "total_display": money(total),
        })
    return invoices


def invoice_export_metadata_json(args: argparse.Namespace) -> None:
    payload = json.load(sys.stdin) if not sys.stdin.isatty() else {}
    book = Path(args.book).expanduser()
    date_from = parse_form_date(str(payload.get("date_from") or ""))
    date_to = parse_form_date(str(payload.get("date_to") or ""))
    customer_ids = dedupe(payload.get("customer_ids", []))
    try:
        con = sqlite_connect_ro(book)
    except sqlite3.Error as exc:
        raise RuntimeError(f"Invoice export requires a readable SQLite GnuCash book: {exc}") from exc
    try:
        invoices = load_invoice_export_data(con, date_from=date_from, date_to=date_to, customer_ids=customer_ids)
    finally:
        con.close()
    print_json({
        "ok": True,
        "invoice_count": len(invoices),
        "invoices": [
            {
                "guid": inv["guid"],
                "id": inv["id"],
                "customer_id": inv["customer_id"],
                "customer_name": inv["customer_name"],
                "billing_name": inv.get("billing_name", ""),
                "date": inv["date"],
                "due_date": inv["due_date"],
                "description": inv["description"],
                "total": str(inv["total"]),
                "total_display": inv["total_display"],
            }
            for inv in invoices
        ],
    })


def render_invoice_filename(invoice: dict[str, Any], payload: dict[str, Any]) -> str:
    customer = invoice.get("customer", {})
    template = str(payload.get("filename_template") or "{customer} - {invoice_id} - invoice")
    source = str(payload.get("filename_customer_source") or "billing_name")
    date_fmt = str(payload.get("filename_date_format") or "Y-m-d")
    inv_date = filename_date(invoice.get("date", ""), date_fmt)
    date_from = filename_date(payload.get("date_from", ""), date_fmt)
    date_to = filename_date(payload.get("date_to", ""), date_fmt)
    selected_customer = filename_customer_value(customer, source)
    values = {
        "customer": selected_customer,
        "customer_id": str(customer.get("id") or invoice.get("customer_id") or ""),
        "customer_number": str(customer.get("id") or invoice.get("customer_id") or ""),
        "company_name": str(customer.get("name") or invoice.get("customer_name") or ""),
        "name": str(customer.get("name") or invoice.get("customer_name") or ""),
        "billing_name": str(customer.get("billing_name") or invoice.get("billing_name") or ""),
        "invoice_id": str(invoice.get("id") or ""),
        "invoice_date": inv_date,
        "date": inv_date,
        "date_from": date_from,
        "date_to": date_to,
        "text": str(payload.get("filename_text") or "invoice"),
    }

    def repl(match: re.Match[str]) -> str:
        return values.get(match.group(1).strip().lower(), "")

    rendered = re.sub(r"\{([A-Za-z0-9_ -]+)\}", repl, template)
    rendered = re.sub(r"\s+", " ", rendered).strip()
    return safe_filename(rendered)


def render_invoice_html(invoice: dict[str, Any], payload: dict[str, Any], style_css: str) -> str:
    page_size = str(payload.get("page_size") or "Letter")
    css = INVOICE_CSS.replace("__PAGE_SIZE__", page_size) + "\n" + style_css + "\n" + str(payload.get("custom_css") or "")
    customer = invoice.get("customer", {})
    logo = ""
    logo_path = str(payload.get("logo_path") or "")
    if logo_path and Path(logo_path).is_file():
        logo = f'<img class="logo" src="{html_escape(Path(logo_path).resolve().as_uri())}" alt="Logo">'
    bill_name = customer.get("billing_name") or customer.get("name") or invoice.get("customer_name") or ""
    entries = invoice.get("entries") or []
    if entries:
        lines = []
        for entry in entries:
            lines.append(
                "<tr>"
                f"<td>{html_escape(entry.get('description',''))}</td>"
                f"<td>{html_escape(entry.get('action',''))}</td>"
                f"<td>{html_escape(entry.get('account',''))}</td>"
                f"<td class='num'>{html_escape(str(entry.get('quantity','')))}</td>"
                f"<td class='num'>{html_escape(money(entry.get('unit_price', Decimal('0'))))}</td>"
                f"<td class='num'>{html_escape(money(entry.get('amount', Decimal('0'))))}</td>"
                "</tr>"
            )
        body = "\n".join(lines)
    else:
        body = f"<tr><td>{html_escape(invoice.get('description','Invoice'))}</td><td></td><td>{html_escape(invoice.get('ar_account',''))}</td><td class='num'>1</td><td class='num'>{html_escape(money(invoice.get('total', Decimal('0'))))}</td><td class='num'>{html_escape(money(invoice.get('total', Decimal('0'))))}</td></tr>"
    footer = f"<div class='footer'>{html_escape(payload.get('footer_text',''))}</div>" if payload.get("footer_text") else ""
    html = f"""<!doctype html><html><head><meta charset="utf-8"><style>{css}</style></head><body>
<div class="invoice-header">
  <div>
    <h1 class="invoice-title">Invoice {html_escape(invoice.get('id',''))}</h1>
    <div><strong>Date:</strong> {html_escape(display_date(invoice.get('date','')))}</div>
    <div><strong>Due date:</strong> {html_escape(display_date(invoice.get('due_date','')))}</div>
    <div><strong>Customer ID:</strong> {html_escape(invoice.get('customer_id',''))}</div>
  </div>
  <div class="invoice-meta">{logo}<div>{html_escape(payload.get('organization_name',''))}</div></div>
</div>
<div class="bill-to"><h2>Bill To</h2>{html_escape(bill_name)}<br>{html_escape(customer.get('name',''))}</div>
<table><thead><tr><th>Description</th><th>Action</th><th>Account</th><th class="num">Qty</th><th class="num">Unit Price</th><th class="num">Amount</th></tr></thead><tbody>{body}<tr class="invoice-total"><td colspan="5" class="num">Total</td><td class="num">{html_escape(money(invoice.get('total', Decimal('0'))))}</td></tr></tbody></table>
<div class="notes">{html_escape(invoice.get('notes',''))}</div>
{footer}
</body></html>"""
    return html


def invoice_pdfs_json(args: argparse.Namespace) -> None:
    payload = json.load(sys.stdin)
    book = Path(args.book).expanduser()
    out_dir = Path(args.out_dir).expanduser()
    html_dir = out_dir / "html"
    pdf_dir = out_dir / "pdf"
    ensure_runtime_dir(html_dir)
    ensure_runtime_dir(pdf_dir)
    invoice_guids = dedupe(payload.get("invoice_guids", []))
    if not invoice_guids:
        raise RuntimeError("No invoice GUIDs were selected for export")
    style_css = extract_style_reference(str(payload.get("style_reference_path") or ""))
    try:
        con = sqlite_connect_ro(book)
    except sqlite3.Error as exc:
        raise RuntimeError(f"Invoice PDF export requires a readable SQLite GnuCash book: {exc}") from exc
    try:
        invoices = load_invoice_export_data(con, invoice_guids=invoice_guids)
    finally:
        con.close()

    generated: list[dict[str, Any]] = []
    used_names: set[str] = set()
    for invoice in invoices:
        html = render_invoice_html(invoice, payload, style_css)
        base = render_invoice_filename(invoice, payload)
        unique_base = base
        n = 2
        while unique_base.lower() in used_names:
            unique_base = safe_filename(f"{base} ({n})")
            n += 1
        used_names.add(unique_base.lower())
        html_path = html_dir / f"{unique_base}.html"
        pdf_path = pdf_dir / f"{unique_base}.pdf"
        html_path.write_text(html, encoding="utf-8")
        apply_runtime_permissions(html_path)
        render_pdf(str(payload.get("chromium_bin") or ""), html_path, pdf_path, show_page_numbers=bool(payload.get("show_page_numbers", True)))
        apply_runtime_permissions(pdf_path)
        generated.append({
            "invoice_id": invoice["id"],
            "invoice_guid": invoice["guid"],
            "customer_id": invoice["customer_id"],
            "customer_name": invoice["customer_name"],
            "html": str(html_path),
            "pdf": str(pdf_path),
            "status": "generated",
        })

    zip_path = out_dir / "invoice-pdfs.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for item in generated:
            zf.write(item["pdf"], arcname=Path(item["pdf"]).name)
    apply_runtime_permissions(zip_path)
    manifest = {
        "ok": True,
        "book": str(book),
        "date_from": str(payload.get("date_from", "")),
        "date_to": str(payload.get("date_to", "")),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "pdf_count": len(generated),
        "zip": str(zip_path),
        "invoices": generated,
    }
    manifest_path = out_dir / "batch-manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    apply_runtime_permissions(manifest_path)
    apply_runtime_permissions(out_dir)
    print_json(manifest)

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="GnuCash batch invoice creation helper")
    sub = parser.add_subparsers(dest="command", required=True)

    p = sub.add_parser("scan-book", help="Scan a GnuCash book for customers and invoice IDs")
    p.add_argument("--book", required=True)
    p.add_argument("--prefix", default="")
    p.add_argument("--padding", default="0")
    p.set_defaults(func=scan_book_json)

    p = sub.add_parser("scan-upload", help="Scan an uploaded customer ID file")
    p.add_argument("--book", required=True)
    p.add_argument("--input", required=True)
    p.add_argument("--prefix", default="")
    p.add_argument("--padding", default="0")
    p.set_defaults(func=scan_upload_json)

    p = sub.add_parser("scan-ids", help="Validate customer IDs from JSON stdin")
    p.add_argument("--book", required=True)
    p.add_argument("--prefix", default="")
    p.add_argument("--padding", default="0")
    p.set_defaults(func=scan_ids_json)

    p = sub.add_parser("generate", help="Generate GnuCash invoice import CSV from JSON stdin")
    p.add_argument("--book", required=True)
    p.add_argument("--out", required=True)
    p.add_argument("--prefix", default="")
    p.add_argument("--padding", default="0")
    p.set_defaults(func=generate_json)

    p = sub.add_parser("report-metadata", help="Scan a SQLite GnuCash book for report metadata such as A/R accounts")
    p.add_argument("--book", required=True)
    p.set_defaults(func=report_metadata_json)

    p = sub.add_parser("customer-reports", help="Generate batch customer report HTML/PDF/ZIP from JSON stdin")
    p.add_argument("--book", required=True)
    p.add_argument("--out-dir", required=True)
    p.set_defaults(func=customer_reports_json)

    p = sub.add_parser("invoice-export-metadata", help="List invoices available for per-invoice PDF export")
    p.add_argument("--book", required=True)
    p.set_defaults(func=invoice_export_metadata_json)

    p = sub.add_parser("invoice-pdfs", help="Generate one invoice PDF per selected invoice GUID from JSON stdin")
    p.add_argument("--book", required=True)
    p.add_argument("--out-dir", required=True)
    p.set_defaults(func=invoice_pdfs_json)

    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        args.func(args)
    except Exception as exc:  # noqa: BLE001
        print_json({"ok": False, "error": str(exc)})
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
